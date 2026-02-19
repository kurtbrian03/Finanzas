from __future__ import annotations

from datetime import datetime
import json
from pathlib import Path
import zipfile

import fitz
from openpyxl.chart import BarChart, Reference
import pandas as pd


def _as_dict(valor: object) -> dict[str, object]:
    return valor if isinstance(valor, dict) else {}


def _safe_name(base: str) -> str:
    return "".join(c if c.isalnum() or c in {"-", "_"} else "_" for c in base)


def generar_reporte_txt(resumen: dict[str, object], out_path: Path) -> Path:
    tipos = _as_dict(resumen.get("tipos", {}))
    lineas = [
        "REPORTE ANALÍTICO DOCUMENTAL",
        f"Fecha: {datetime.now().isoformat(timespec='seconds')}",
        "",
        f"Total archivos: {resumen.get('total_archivos', 0)}",
        f"Total bytes: {resumen.get('total_bytes', 0)}",
        f"PDF páginas: {resumen.get('pdf_paginas', 0)}",
        f"Excel hojas: {resumen.get('excel_hojas', 0)}",
        f"Imágenes con logo probable: {resumen.get('imagenes_logo_probable', 0)}",
        f"Palabras en texto: {resumen.get('texto_palabras', 0)}",
        f"Entradas ZIP: {resumen.get('zip_entradas', 0)}",
        "",
        "Conteo por tipo:",
    ]
    for tipo, cantidad in tipos.items():
        lineas.append(f"- {tipo}: {cantidad}")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lineas) + "\n", encoding="utf-8")
    return out_path


def anexar_resumen_auditoria_txt(out_path: Path, auditoria_data: dict[str, object]) -> Path:
    """Anexa resumen de auditoría de búsqueda al reporte TXT."""
    if not out_path.exists():
        return out_path
    metadata = _as_dict(auditoria_data.get("metadata", {}))
    score_summary = _as_dict(auditoria_data.get("resumen_scores", {}))
    query_ctx = _as_dict(auditoria_data.get("query_context", {}))
    lineas = [
        "",
        "Auditoría de búsqueda:",
        f"- Engine version: {metadata.get('engine_version', '')}",
        f"- Generated at: {metadata.get('generated_at', '')}",
        f"- Modo: {query_ctx.get('modo', '')}",
        f"- Query: {query_ctx.get('query_raw', '')}",
        f"- Result rows: {metadata.get('result_rows', 0)}",
        f"- Score final avg: {score_summary.get('score_final_avg', 0)}",
        f"- Score boosting avg: {score_summary.get('score_boosting_avg', 0)}",
    ]
    out_path.write_text(out_path.read_text(encoding="utf-8") + "\n".join(lineas) + "\n", encoding="utf-8")
    return out_path


def anexar_resumen_performance_txt(out_path: Path, auditoria_data: dict[str, object]) -> Path:
    """Anexa resumen de performance (profiling) al reporte TXT cuando existe."""
    if not out_path.exists():
        return out_path
    perf = _as_dict(auditoria_data.get("performance_metrics", {}))
    if not perf:
        return out_path

    components = _as_dict(perf.get("components", {}))
    total_time_ms = perf.get("total_time_ms", 0)
    lineas = [
        "",
        "Performance del motor (profiling):",
        f"- Timestamp: {perf.get('timestamp', '')}",
        f"- Versión motor: {perf.get('version_motor', '')}",
        f"- Tiempo total (ms): {total_time_ms}",
        "- Componentes:",
    ]
    for component, value in components.items():
        lineas.append(f"  - {component}: {value} ms")

    out_path.write_text(out_path.read_text(encoding="utf-8") + "\n".join(lineas) + "\n", encoding="utf-8")
    return out_path


def generar_reporte_excel(
    resumen: dict[str, object],
    analitica_detallada: list[dict[str, object]],
    tree_summary: dict[str, object],
    out_path: Path,
    auditoria_data: dict[str, object] | None = None,
) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tipos = _as_dict(resumen.get("tipos", {}))
    proveedores_top = _as_dict(tree_summary.get("proveedores_top", {}))
    hospitales_top = _as_dict(tree_summary.get("hospitales_top", {}))
    meses_top = _as_dict(tree_summary.get("meses_top", {}))

    df_resumen = pd.DataFrame(
        [
            {"indicador": "total_archivos", "valor": resumen.get("total_archivos", 0)},
            {"indicador": "total_bytes", "valor": resumen.get("total_bytes", 0)},
            {"indicador": "pdf_paginas", "valor": resumen.get("pdf_paginas", 0)},
            {"indicador": "excel_hojas", "valor": resumen.get("excel_hojas", 0)},
            {"indicador": "imagenes_logo_probable", "valor": resumen.get("imagenes_logo_probable", 0)},
            {"indicador": "texto_palabras", "valor": resumen.get("texto_palabras", 0)},
            {"indicador": "zip_entradas", "valor": resumen.get("zip_entradas", 0)},
        ]
    )
    df_tipos = pd.DataFrame(
        [{"tipo": k, "cantidad": v} for k, v in tipos.items()]
    )
    df_detalle = pd.DataFrame(analitica_detallada)
    df_tree = pd.DataFrame(
        [
            {"grupo": "proveedor", "clave": k, "cantidad": v}
            for k, v in proveedores_top.items()
        ]
        + [
            {"grupo": "hospital", "clave": k, "cantidad": v}
            for k, v in hospitales_top.items()
        ]
        + [
            {"grupo": "mes", "clave": k, "cantidad": v}
            for k, v in meses_top.items()
        ]
    )
    df_auditoria = pd.DataFrame()
    df_performance = pd.DataFrame()
    if isinstance(auditoria_data, dict):
        resultados_scores = auditoria_data.get("resultados_scores", [])
        if isinstance(resultados_scores, list):
            df_auditoria = pd.DataFrame(resultados_scores)
        perf = _as_dict(auditoria_data.get("performance_metrics", {}))
        components = _as_dict(perf.get("components", {}))
        if components:
            df_performance = pd.DataFrame(
                [
                    {"component": k, "time_ms": v}
                    for k, v in components.items()
                ]
            )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
        df_tipos.to_excel(writer, sheet_name="Tipos", index=False)
        df_detalle.to_excel(writer, sheet_name="Detalle", index=False)
        df_tree.to_excel(writer, sheet_name="ArbolVirtual", index=False)
        if not df_auditoria.empty:
            cols = [
                c
                for c in [
                    "ruta",
                    "score_exacto",
                    "score_fuzzy",
                    "score_semantico",
                    "score_tokens",
                    "score_temporal",
                    "score_estructural",
                    "score_boosting",
                    "score_final",
                    "relevancia",
                ]
                if c in df_auditoria.columns
            ]
            if cols:
                df_auditoria[cols].to_excel(writer, sheet_name="AuditoriaSearch", index=False)
        if not df_performance.empty:
            df_performance.to_excel(writer, sheet_name="PerformanceSearch", index=False)

        wb = writer.book
        ws = wb["Tipos"]
        if ws.max_row >= 2:
            chart = BarChart()
            chart.title = "Archivos por tipo"
            chart.y_axis.title = "Cantidad"
            chart.x_axis.title = "Tipo"
            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 7
            chart.width = 12
            chart.anchor = "E2"
            ws.add_chart(chart)

    return out_path


def generar_reporte_pdf(resumen: dict[str, object], tree_summary: dict[str, object], out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    doc = fitz.open()
    pagina = doc.new_page(width=595, height=842)

    tipos = _as_dict(resumen.get("tipos", {}))
    proveedores_top = _as_dict(tree_summary.get("proveedores_top", {}))

    lineas = [
        "REPORTE ANALÍTICO DOCUMENTAL",
        f"Generado: {datetime.now().isoformat(timespec='seconds')}",
        "",
        f"Total archivos: {resumen.get('total_archivos', 0)}",
        f"Total bytes: {resumen.get('total_bytes', 0)}",
        f"PDF páginas: {resumen.get('pdf_paginas', 0)}",
        f"Excel hojas: {resumen.get('excel_hojas', 0)}",
        f"Imágenes logo probable: {resumen.get('imagenes_logo_probable', 0)}",
        f"Palabras texto: {resumen.get('texto_palabras', 0)}",
        f"Entradas ZIP: {resumen.get('zip_entradas', 0)}",
        "",
        "Tipos:",
    ]
    for tipo, cantidad in tipos.items():
        lineas.append(f"- {tipo}: {cantidad}")

    lineas.append("")
    lineas.append("Top proveedores:")
    for clave, cantidad in proveedores_top.items():
        lineas.append(f"- {clave}: {cantidad}")

    texto = "\n".join(lineas)
    pagina.insert_text((42, 42), texto, fontsize=10)
    doc.save(out_path)
    doc.close()
    return out_path


def generar_zip_reportes(archivos: list[Path], out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for archivo in archivos:
            if archivo.exists():
                zf.write(archivo, arcname=archivo.name)
    return out_path


def exportar_json(data: dict[str, object], out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path


def generar_paquete_reportes(
    resumen: dict[str, object],
    analitica_detallada: list[dict[str, object]],
    tree_summary: dict[str, object],
    out_dir: str | Path,
    prefijo: str = "dropbox_analytics",
    auditoria_data: dict[str, object] | None = None,
) -> dict[str, str]:
    out = Path(out_dir)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = _safe_name(f"{prefijo}_{ts}")

    txt_path = generar_reporte_txt(resumen, out / f"{base}.txt")
    excel_path = generar_reporte_excel(
        resumen,
        analitica_detallada,
        tree_summary,
        out / f"{base}.xlsx",
        auditoria_data=auditoria_data,
    )
    pdf_path = generar_reporte_pdf(resumen, tree_summary, out / f"{base}.pdf")
    archivos_zip = [txt_path, excel_path, pdf_path]

    audit_json_path: Path | None = None
    audit_csv_path: Path | None = None
    performance_json_path: Path | None = None
    performance_csv_path: Path | None = None
    if isinstance(auditoria_data, dict):
        audit_json_path = exportar_json(auditoria_data, out / f"{base}_audit.json")
        resultados_scores = auditoria_data.get("resultados_scores", [])
        if isinstance(resultados_scores, list):
            df_audit = pd.DataFrame(resultados_scores)
            cols = [
                c
                for c in [
                    "ruta",
                    "score_exacto",
                    "score_fuzzy",
                    "score_semantico",
                    "score_tokens",
                    "score_temporal",
                    "score_estructural",
                    "score_boosting",
                    "score_final",
                ]
                if c in df_audit.columns
            ]
            if cols:
                audit_csv_path = out / f"{base}_audit.csv"
                df_audit[cols].to_csv(audit_csv_path, index=False, encoding="utf-8")

        perf = _as_dict(auditoria_data.get("performance_metrics", {}))
        components = _as_dict(perf.get("components", {}))
        if perf:
            performance_json_path = exportar_json(perf, out / f"{base}_performance.json")
        if components:
            performance_csv_path = out / f"{base}_performance.csv"
            pd.DataFrame(
                [{"component": k, "time_ms": v} for k, v in components.items()]
            ).to_csv(performance_csv_path, index=False, encoding="utf-8")

        anexar_resumen_auditoria_txt(txt_path, auditoria_data)
        anexar_resumen_performance_txt(txt_path, auditoria_data)

    if audit_json_path is not None:
        archivos_zip.append(audit_json_path)
    if audit_csv_path is not None and audit_csv_path.exists():
        archivos_zip.append(audit_csv_path)
    if performance_json_path is not None and performance_json_path.exists():
        archivos_zip.append(performance_json_path)
    if performance_csv_path is not None and performance_csv_path.exists():
        archivos_zip.append(performance_csv_path)

    zip_path = generar_zip_reportes(archivos_zip, out / f"{base}.zip")

    salida = {
        "txt": str(txt_path),
        "excel": str(excel_path),
        "pdf": str(pdf_path),
        "zip": str(zip_path),
    }
    if audit_json_path is not None:
        salida["audit_json"] = str(audit_json_path)
    if audit_csv_path is not None and audit_csv_path.exists():
        salida["audit_csv"] = str(audit_csv_path)
    if performance_json_path is not None and performance_json_path.exists():
        salida["performance_json"] = str(performance_json_path)
    if performance_csv_path is not None and performance_csv_path.exists():
        salida["performance_csv"] = str(performance_csv_path)
    return salida
